import os
from datetime import datetime
import pytz

from email_validator import EmailNotValidError, validate_email
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "mysite/newsletter_emails.xlsx"

# Create workbook if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws.append(["Email", "Date Subscribed"])
    wb.save(EXCEL_FILE)


def is_email_in_excel(email):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == email:
            return True
    return False


def add_email_to_excel(email):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    bangladesh_tz = pytz.timezone("Asia/Dhaka")
    current_date = datetime.now(bangladesh_tz).strftime("%d-%m-%Y | %H:%M:%S")

    ws.append([email, current_date])
    wb.save(EXCEL_FILE)


@app.route("/subscribe", methods=["POST"])
def subscribe():
    try:
        data = request.get_json()

        if not data or "email" not in data:
            return jsonify({"error": "No email provided"}), 400

        email = data["email"]

        try:
            validated_email = validate_email(email).email
        except EmailNotValidError as e:
            return jsonify({"error": str(e)}), 400

        if is_email_in_excel(validated_email):
            return jsonify({"error": "Email already subscribed"}), 400

        add_email_to_excel(validated_email)

        return jsonify({"message": "Successfully subscribed"}), 200

    except Exception as e:
        print(str(e))
        return jsonify({"error": "Something went wrong", "details": str(e)}), 500

@app.route("/subscribers", methods=["GET"])
def get_subscribers():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        subscribers = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            email, date_subscribed = row
            subscribers.append({
                "email": email,
                "date_subscribed": date_subscribed
            })

        return jsonify(subscribers), 200

    except Exception as e:
        print(str(e))
        return jsonify({"error": "Failed to fetch subscribers", "details": str(e)}), 500

@app.route("/delete/<password>/<email>", methods=["GET"])
def delete_email(password, email):
    try:
        # Set your own secret password here
        SECRET_PASSWORD = ""

        if password != SECRET_PASSWORD:
            return jsonify({"error": "Invalid password"}), 403

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        found = False
        for row in range(2, ws.max_row + 1):  # Skip header
            cell_email = ws[f"A{row}"].value
            if cell_email == email:
                ws.delete_rows(row)
                wb.save(EXCEL_FILE)
                found = True
                break

        if not found:
            return jsonify({"error": "Email not found"}), 404

        return jsonify({"message": f"Email '{email}' deleted successfully"}), 200

    except Exception as e:
        print(str(e))
        return jsonify({"error": "Something went wrong", "details": str(e)}), 500

@app.route("/", methods=["GET"])
def index():
    return jsonify({"message": "Hello, World from Newsletter!"}), 200
